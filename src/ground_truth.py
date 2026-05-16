"""Load ground_truth.xlsx → lookup dict keyed by (real_filename, date).

PHI NOTE: This module handles REAL filenames from ground_truth.xlsx.
The returned dict is for internal use only — passed to evaluation.py
where it is used for GT matching. Real filenames are NEVER written to
any output artifact.

Ground truth sheet: "Timesheet Extract"
No header row. 6 positional columns:
    [0] source_file   — real filename (e.g. "N.Rivera-Timesheets-021826-022426.pdf")
    [1] date          — ISO string (e.g. "2026-02-18")
    [2] total_hours   — float (e.g. 8.0)
    [3] time_in       — 12-hour string (e.g. "7:00 AM")
    [4] time_out      — 12-hour string (e.g. "3:00 PM")
    [5] employee_name — real name (NOT evaluated — ignored)
"""

from __future__ import annotations

import datetime
import logging
from pathlib import Path

import openpyxl

from src.models import GroundTruthRow

logger = logging.getLogger(__name__)

GT_SHEET_NAME = "Timesheet Extract"

# Column indices (zero-based, no header row)
_COL_SOURCE_FILE = 0
_COL_DATE = 1
_COL_TOTAL_HOURS = 2
_COL_TIME_IN = 3
_COL_TIME_OUT = 4
# _COL_EMPLOYEE_NAME = 5  ← not evaluated


def load_ground_truth(
    gt_path: Path,
) -> dict[tuple[str, datetime.date], GroundTruthRow]:
    """Read ground_truth.xlsx and return an O(1) lookup dict.

    Args:
        gt_path: Path to ground_truth.xlsx.

    Returns:
        Dict keyed by (real_source_filename, date) → GroundTruthRow.
        Real filenames are NEVER written to output — internal use only.

    Raises:
        FileNotFoundError: If gt_path does not exist.
    """
    if not gt_path.exists():
        raise FileNotFoundError(f"Ground truth file not found: {gt_path}")

    wb = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)

    if GT_SHEET_NAME not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        raise KeyError(
            f"Sheet '{GT_SHEET_NAME}' not found in {gt_path}. "
            f"Available sheets: {available}"
        )

    ws = wb[GT_SHEET_NAME]
    lookup: dict[tuple[str, datetime.date], GroundTruthRow] = {}
    skipped = 0
    duplicates = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip completely empty rows
        if not any(c is not None for c in row):
            continue

        if len(row) < 5:
            skipped += 1
            logger.debug("GT row %d has fewer than 5 columns — skipped", row_idx)
            continue

        source_file = row[_COL_SOURCE_FILE]
        date_raw = row[_COL_DATE]
        total_hours_raw = row[_COL_TOTAL_HOURS]
        time_in_raw = row[_COL_TIME_IN]
        time_out_raw = row[_COL_TIME_OUT]

        if not source_file or not date_raw:
            skipped += 1
            continue

        # Parse date
        date_obj = _parse_date(date_raw, row_idx)
        if date_obj is None:
            skipped += 1
            continue

        # Parse times (12-hour format "7:00 AM")
        time_in = _parse_12h_time(time_in_raw, row_idx, "time_in") if time_in_raw else None
        time_out = _parse_12h_time(time_out_raw, row_idx, "time_out") if time_out_raw else None

        # Parse hours
        total_hours: float | None = None
        if total_hours_raw is not None:
            try:
                total_hours = float(total_hours_raw)
            except (TypeError, ValueError):
                logger.debug("GT row %d — non-numeric total_hours: %r", row_idx, total_hours_raw)

        key = (str(source_file).strip(), date_obj)

        if key in lookup:
            duplicates += 1
            logger.debug(
                "GT row %d — duplicate key %s / %s (keeping first)",
                row_idx, source_file, date_obj,
            )
            continue

        lookup[key] = GroundTruthRow(
            source_file=str(source_file).strip(),
            date=date_obj,
            time_in=time_in,
            time_out=time_out,
            total_hours=total_hours,
        )

    wb.close()
    logger.info(
        "Loaded %d GT entries (%d skipped, %d duplicates) from %s",
        len(lookup), skipped, duplicates, gt_path.name,
    )
    return lookup


def _parse_date(raw: object, row_idx: int) -> datetime.date | None:
    """Parse ISO string or datetime.date/datetime.datetime → datetime.date."""
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    try:
        return datetime.date.fromisoformat(str(raw).strip())
    except ValueError:
        logger.debug("GT row %d — cannot parse date: %r", row_idx, raw)
        return None


def _parse_12h_time(raw: object, row_idx: int, field: str) -> datetime.time | None:
    """Parse 12-hour time string like '7:00 AM' or '11:30 PM' → datetime.time."""
    if isinstance(raw, datetime.time):
        return raw
    s = str(raw).strip().upper()   # uppercase the VALUE for %p matching
    for fmt in ("%I:%M %p", "%I:%M%p"):
        try:
            return datetime.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    # Fallback: 24h format in case GT has mixed data
    try:
        return datetime.time.fromisoformat(str(raw).strip())
    except ValueError:
        pass
    logger.debug("GT row %d — cannot parse %s time: %r", row_idx, field, raw)
    return None
