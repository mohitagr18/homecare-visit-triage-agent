"""Ingest merged_results.xlsx → list[ExtractionRow].

Reads the "Timesheet Data" sheet from a method's merged_results.xlsx.
All source_file values remain anonymized (e.g., "patient_l_week5.pdf").

Column mapping (case-insensitive header match):
    Source File      → source_file
    Date             → date
    Time In          → time_in
    Time Out         → time_out
    Total Hours      → total_hours
    Calculated Hours → calculated_hours
    Status           → status
    Issues           → issues
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from src.models import ExtractionRow

logger = logging.getLogger(__name__)

# Map of normalized header → ExtractionRow field name
_HEADER_MAP = {
    "source file": "source_file",
    "date": "date",
    "time in": "time_in",
    "time out": "time_out",
    "total hours": "total_hours",
    "calculated hours": "calculated_hours",
    "status": "status",
    "issues": "issues",
}

_REQUIRED_FIELDS = {"source_file", "date", "time_in", "time_out", "total_hours", "status"}
SHEET_NAME = "Timesheet Data"


def ingest(merged_path: Path) -> list[ExtractionRow]:
    """Read merged_results.xlsx and return a list of ExtractionRow objects.

    Args:
        merged_path: Path to merged_results.xlsx for a single method.

    Returns:
        List of ExtractionRow objects. Rows with missing required fields are skipped.

    Raises:
        FileNotFoundError: If merged_path does not exist.
        KeyError: If the required sheet name is not found.
    """
    if not merged_path.exists():
        raise FileNotFoundError(f"merged_results.xlsx not found: {merged_path}")

    wb = openpyxl.load_workbook(merged_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        raise KeyError(
            f"Sheet '{SHEET_NAME}' not found in {merged_path}. "
            f"Available sheets: {available}"
        )

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)

    # Parse header row
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        wb.close()
        logger.warning("Empty sheet in %s", merged_path)
        return []

    col_map: dict[str, int] = {}  # field_name → column index
    for idx, cell in enumerate(header_raw):
        if cell is None:
            continue
        normalized = str(cell).strip().lower()
        if normalized in _HEADER_MAP:
            col_map[_HEADER_MAP[normalized]] = idx

    missing_required = _REQUIRED_FIELDS - set(col_map.keys())
    if missing_required:
        wb.close()
        raise KeyError(
            f"Required columns missing from {merged_path}: {missing_required}. "
            f"Found columns: {list(col_map.keys())}"
        )

    results: list[ExtractionRow] = []
    skipped = 0

    for row_idx, row in enumerate(rows_iter, start=2):  # 1-indexed, row 1 was header

        def _get(field: str) -> str | float | None:
            idx = col_map.get(field)
            if idx is None:
                return None
            val = row[idx]
            return val

        source_file = _get("source_file")
        date_val = _get("date")
        time_in = _get("time_in")
        time_out = _get("time_out")
        total_hours = _get("total_hours")
        status = _get("status")

        # Skip rows where required fields are None/empty
        if not all([source_file, date_val, time_in, time_out, total_hours is not None, status]):
            skipped += 1
            logger.debug("Row %d skipped — missing required fields", row_idx)
            continue

        # Coerce types
        try:
            total_hours_f = float(total_hours)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            skipped += 1
            logger.debug("Row %d skipped — non-numeric total_hours: %r", row_idx, total_hours)
            continue

        calc_hours = _get("calculated_hours")
        calc_hours_f: float | None = None
        if calc_hours is not None:
            try:
                calc_hours_f = float(calc_hours)
            except (TypeError, ValueError):
                pass

        results.append(
            ExtractionRow(
                source_file=str(source_file).strip(),
                row_index=row_idx,
                date=str(date_val).strip(),
                time_in=str(time_in).strip(),
                time_out=str(time_out).strip(),
                total_hours=total_hours_f,
                calculated_hours=calc_hours_f,
                status=str(status).strip().lower(),
                issues=str(_get("issues")).strip() if _get("issues") else None,
            )
        )

    wb.close()
    logger.info(
        "Ingested %d rows from %s (%d skipped)",
        len(results), merged_path.name, skipped,
    )
    return results
